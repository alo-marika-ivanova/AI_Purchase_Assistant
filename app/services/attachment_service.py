from __future__ import annotations

import csv
import hashlib
import io
import mimetypes
import os
import uuid
from pathlib import Path

from openpyxl import load_workbook

from app.db.repository import PurchasingRepository
from app.services.simple_chat_service import send_or_display_outbound_message

repo = PurchasingRepository()

PROJECT_ROOT = Path(__file__).resolve().parents[2]

ATTACHMENT_STORAGE_DIR = Path(
    os.getenv("ATTACHMENT_STORAGE_DIR", str(PROJECT_ROOT / "data" / "attachments"))
)


def save_case_attachment(
    case_id: int,
    original_filename: str,
    file_bytes: bytes,
    supplier_id: int | None = None,
    message_id: int | None = None,
    channel: str = "manual",
    direction: str = "outbound",
) -> dict:
    """Persist one uploaded file to disk and record its metadata.

    This is the single storage path shared by the UI, email worker, and
    WhatsApp webhook (later phases), so every attachment - regardless of
    channel - ends up hashed, stored under its case, and recorded the same
    way.
    """
    case_dir = ATTACHMENT_STORAGE_DIR / str(case_id)
    case_dir.mkdir(parents=True, exist_ok=True)

    suffix = Path(original_filename).suffix
    stored_path = case_dir / f"{uuid.uuid4().hex}{suffix}"
    stored_path.write_bytes(file_bytes)

    sha256_hash = hashlib.sha256(file_bytes).hexdigest()
    mime_type, _ = mimetypes.guess_type(original_filename)

    attachment_id = repo.add_attachment(
        case_id=case_id,
        original_filename=original_filename,
        stored_path=str(stored_path),
        mime_type=mime_type,
        size_bytes=len(file_bytes),
        sha256_hash=sha256_hash,
        supplier_id=supplier_id,
        message_id=message_id,
        channel=channel,
        direction=direction,
    )

    return {
        "id": attachment_id,
        "original_filename": original_filename,
        "stored_path": str(stored_path),
        "mime_type": mime_type,
        "size_bytes": len(file_bytes),
        "sha256_hash": sha256_hash,
    }


def list_case_attachments(case_id: int) -> list[dict]:
    return repo.list_attachments_for_case(case_id)


def send_case_attachment(
    case_id: int,
    supplier_id: int,
    attachment_id: int,
    caption: str | None = None,
) -> dict:
    """Send an already-uploaded case attachment to one supplier.

    Reuses send_or_display_outbound_message for the actual delivery, so a
    real-communication case with an email-channel supplier gets the file as
    a real SMTP attachment (through the same outbox/retry path as any other
    message), while a simulated case - or a WhatsApp-channel supplier, whose
    integration only supports text today - just records it.
    """
    attachment = repo.get_attachment_by_id(attachment_id)
    if attachment is None or int(attachment["case_id"]) != case_id:
        raise ValueError("Attachment not found for this case.")

    body = (caption or "").strip() or (
        f"Please find attached: {attachment['original_filename']}."
    )

    return send_or_display_outbound_message(
        case_id=case_id,
        supplier_id=supplier_id,
        body=body,
        message_type="attachment_share",
        attachment_ids=[attachment_id],
    )


def extract_text_from_spreadsheet(file_bytes: bytes, filename: str) -> str:
    """Dump every non-empty cell of an uploaded CSV/XLSX file as plain text
    lines, so a supplier's price-filled file can be fed into the same
    free-text reply pipeline (the LLM classifier) used for typed messages,
    without changing that classifier or its extraction logic at all.
    """
    lower_name = filename.lower()

    if lower_name.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        lines = [
            " | ".join(cell.strip() for cell in row if cell.strip())
            for row in csv.reader(io.StringIO(text))
        ]
        return "\n".join(line for line in lines if line)

    if lower_name.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        lines = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows():
                cells = [
                    str(cell.value).strip()
                    for cell in row
                    if cell.value is not None and str(cell.value).strip()
                ]
                if cells:
                    lines.append(" | ".join(cells))

        return "\n".join(lines)

    raise ValueError(
        "Unsupported file type for a supplier reply upload. "
        "Use .csv, .xlsx, or .xlsm."
    )
