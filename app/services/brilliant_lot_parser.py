from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

from app.services.spreadsheet_helpers import (
    clean_text,
    coerce_float,
    find_column_containing,
    find_exact_column,
    find_header_row,
)

# Small reusable pure parser for the brilliant-lot RFQ spreadsheet layout
# (see tests/fixtures/rfq_samples/brilianty.xlsx): several component
# "Sizes" rows share one merged "ORDER CTS TOTAL" cell and one merged
# "Price USD/ct" cell - that merged-cell span IS the lot boundary. This
# module is the single definition of what a "lot" is, used both when
# detecting an outgoing RFQ (rfq_detection_service.py) and when
# deterministically pricing a returned, filled-in copy of the same file
# (attachment_service.py). Nothing here touches the database.

_CT_VALUE_PATTERN = re.compile(r"(\d+[.,]?\d*)\s*ct", re.IGNORECASE)
_MM_PRESENCE_PATTERN = re.compile(r"\bmm\b", re.IGNORECASE)
_CT_PRESENCE_PATTERN = re.compile(r"\bct\b", re.IGNORECASE)
_NUMBER_PATTERN = re.compile(r"(\d+[.,]\d+)")


@dataclass(frozen=True)
class BrilliantLotRow:
    size_text: str
    order_cts: float


@dataclass(frozen=True)
class BrilliantLot:
    """One priced lot: several component size rows sharing one merged
    ORDER CTS TOTAL cell and one merged Price USD/ct cell."""

    lot_index: int
    row_start: int
    row_end: int
    rows: list[BrilliantLotRow]
    order_cts_total: float | None
    price_usd_per_ct: float | None
    label: str
    # 'up_to_1ct' / '1ct_and_up' / None when the component rows imply
    # inconsistent supplier categories - never guessed.
    size_bucket: str | None


@dataclass(frozen=True)
class BrilliantSheetParse:
    header_row_index: int
    lots: list[BrilliantLot]


def classify_brilliant_size(size_text: str) -> str | None:
    """Bucket one "Sizes" cell into 'up_to_1ct' or '1ct_and_up'.

    mm-labeled rows are melee (a round brilliant only reaches ~1ct around
    6.5mm diameter, far above anything seen sized in mm in practice), so
    they are always up_to_1ct without needing a precise mm-to-carat
    conversion. ct-labeled rows use a literal >= 1.0 threshold.
    """
    ct_match = _CT_VALUE_PATTERN.search(size_text)
    if ct_match:
        value = float(ct_match.group(1).replace(",", "."))
        return "1ct_and_up" if value >= 1.0 else "up_to_1ct"

    if _MM_PRESENCE_PATTERN.search(size_text):
        return "up_to_1ct"

    return None


def _lot_size_bucket(rows: list[BrilliantLotRow]) -> str | None:
    buckets = {classify_brilliant_size(row.size_text) for row in rows}
    buckets.discard(None)
    if len(buckets) == 1:
        return next(iter(buckets))
    return None


def _unit_bounds(
    rows: list[BrilliantLotRow], presence_pattern: re.Pattern
) -> tuple[str, str] | None:
    """First/last numeric value (in row order) among rows matching one unit
    (mm or ct), e.g. ("0.66", "1.25") - used to build a human-readable
    range label without guessing an mm<->ct conversion."""
    unit_rows = [row for row in rows if presence_pattern.search(row.size_text)]
    if not unit_rows:
        return None

    first_numbers = _NUMBER_PATTERN.findall(unit_rows[0].size_text)
    last_numbers = _NUMBER_PATTERN.findall(unit_rows[-1].size_text)
    if not first_numbers or not last_numbers:
        return None

    return first_numbers[0].replace(",", "."), last_numbers[-1].replace(",", ".")


def _build_lot_label(lot_index: int, rows: list[BrilliantLotRow]) -> str:
    mm_bounds = _unit_bounds(rows, _MM_PRESENCE_PATTERN)
    ct_bounds = _unit_bounds(rows, _CT_PRESENCE_PATTERN)

    parts = []
    if mm_bounds:
        parts.append(f"{mm_bounds[0]}–{mm_bounds[1]} mm")
    if ct_bounds:
        parts.append(f"{ct_bounds[0]}–{ct_bounds[1]} ct")

    range_text = " / ".join(parts) if parts else "unspecified sizes"
    return f"Brilliant lot {lot_index}: {range_text}"


def build_lot_description(rows: list[BrilliantLotRow], order_cts_total: float | None) -> str:
    """Human-readable summary of a lot's component size rows, stored as the
    case_item's source_description so the original per-size breakdown
    behind the lot's single quantity/price stays visible to the buyer."""
    row_texts = "; ".join(f"{row.size_text}: {row.order_cts:g} ct" for row in rows)
    total_text = f" (total {order_cts_total:g} ct)" if order_cts_total else ""
    return f"{len(rows)} size(s){total_text}: {row_texts}"


def parse_brilliant_lots(ws) -> BrilliantSheetParse | None:
    """Recognize a brilliant-lot worksheet and split it into lots.

    Returns None when the sheet doesn't have the expected "Sizes" /
    "ORDER CTS" / "ORDER CTS TOTAL" / "Price USD/ct" columns, or when the
    ORDER CTS TOTAL and Price USD/ct merged-cell ranges don't line up
    row-for-row - an ambiguous shape is left unrecognized rather than
    guessed at.
    """
    header = find_header_row(ws, {"sizes"})
    if header is None:
        return None

    header_row_index, header_cols = header
    sizes_col = header_cols["sizes"]
    order_cts_col = find_exact_column(
        ws, header_row_index, "order cts"
    ) or find_column_containing(ws, header_row_index, "order cts")
    total_col = find_exact_column(
        ws, header_row_index, "order cts total"
    ) or find_column_containing(ws, header_row_index, "total")
    price_col = find_column_containing(ws, header_row_index, "price")

    if order_cts_col is None or total_col is None or price_col is None:
        return None
    if len({sizes_col, order_cts_col, total_col, price_col}) != 4:
        return None

    total_ranges = sorted(
        (rng.min_row, rng.max_row)
        for rng in ws.merged_cells.ranges
        if rng.min_col == total_col
        and rng.max_col == total_col
        and rng.min_row > header_row_index
    )
    price_ranges = sorted(
        (rng.min_row, rng.max_row)
        for rng in ws.merged_cells.ranges
        if rng.min_col == price_col
        and rng.max_col == price_col
        and rng.min_row > header_row_index
    )

    if not total_ranges or total_ranges != price_ranges:
        return None

    lots: list[BrilliantLot] = []

    for lot_index, (row_start, row_end) in enumerate(total_ranges, start=1):
        rows: list[BrilliantLotRow] = []
        for row_index in range(row_start, row_end + 1):
            size_text = clean_text(ws.cell(row=row_index, column=sizes_col).value)
            if not size_text:
                continue

            quantity = coerce_float(ws.cell(row=row_index, column=order_cts_col).value)
            if not quantity:
                continue

            rows.append(BrilliantLotRow(size_text=size_text, order_cts=quantity))

        if not rows:
            continue

        order_cts_total = coerce_float(ws.cell(row=row_start, column=total_col).value)
        price_usd_per_ct = coerce_float(ws.cell(row=row_start, column=price_col).value)

        lots.append(
            BrilliantLot(
                lot_index=lot_index,
                row_start=row_start,
                row_end=row_end,
                rows=rows,
                order_cts_total=order_cts_total,
                price_usd_per_ct=price_usd_per_ct,
                label=_build_lot_label(lot_index, rows),
                size_bucket=_lot_size_bucket(rows),
            )
        )

    if not lots:
        return None

    return BrilliantSheetParse(header_row_index=header_row_index, lots=lots)


def try_build_brilliant_lot_offer_text(file_bytes: bytes, filename: str) -> str | None:
    """Deterministically turn a returned, filled-in copy of the
    brilliant-lot RFQ template into the same pipe-delimited
    "Description | Price" shape produced by
    attachment_service.extract_text_from_spreadsheet for the natural-stone
    template, keyed by each lot's stable label - so the existing
    deterministic multi-item safeguard (rfq_multi_item_price_safeguard.py),
    which matches purely by exact item_material text, can price a
    brilliant-lot reply without any change of its own.

    Returns None - falling back to the generic flattened dump, and from
    there to the LLM - unless the workbook is recognized as a brilliant-lot
    layout AND every lot in it has exactly one clean, positive price. A
    partial or ambiguous file must not be guessed at.
    """
    lower_name = filename.lower()
    if not lower_name.endswith((".xlsx", ".xlsm")):
        return None

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        return None

    for sheet_name in workbook.sheetnames:
        parsed = parse_brilliant_lots(workbook[sheet_name])
        if parsed is None:
            continue

        lines = ["Description | Price USD/ct"]
        for lot in parsed.lots:
            price = lot.price_usd_per_ct
            if price is None or price <= 0:
                return None
            lines.append(f"{lot.label} | {price}")

        return "\n".join(lines)

    return None
