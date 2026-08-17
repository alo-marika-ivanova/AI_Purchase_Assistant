from __future__ import annotations

import difflib
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook

from app.db.repository import PurchasingRepository
from app.services.brilliant_lot_parser import build_lot_description, parse_brilliant_lots
from app.services.spreadsheet_helpers import (
    clean_text,
    coerce_float,
    find_column_containing,
    find_header_row,
)

repo = PurchasingRepository()

NATURAL_STONE_GOODS_GROUP = "Precious stones"
BRILLIANT_GOODS_GROUP = "Diamonds"

# Below this fuzzy-match score, a description is treated as unresolved
# rather than guessed. Calibrated against the real prirodni.xlsx fixture
# (matches score 0.90-1.00) versus unrelated stone names (score <= 0.76).
NATURAL_STONE_MATCH_THRESHOLD = 0.82

_BRILLIANT_COLOR_KEYWORDS = ("blue", "green", "orange", "pink", "yellow")


@dataclass(frozen=True)
class DetectedRfqItem:
    """One catalog item (a natural-stone row, or a brilliant lot) found in
    an uploaded RFQ file, plus enough context for the buyer to review it."""

    goods_name: str
    description: str
    quantity: float | None
    item_label: str | None = None

    @property
    def display_name(self) -> str:
        """Buyer-facing label, also used as the created subcase's
        item_material - distinct per row for natural stones (whose catalog
        goods_name alone doesn't capture size/shape) and distinct per lot
        for brilliants (several lots can share the same catalog bucket but
        must still become separate case_items, since each lot is priced
        independently)."""
        return self.item_label or self.goods_name


@dataclass(frozen=True)
class RfqDetectionResult:
    """Outcome of trying to read case setup (item(s) + implied suppliers)
    directly out of an uploaded RFQ spreadsheet."""

    recognized: bool
    file_type: str | None = None
    items: list[DetectedRfqItem] = field(default_factory=list)
    unresolved_lines: list[str] = field(default_factory=list)


def detect_rfq_selection(file_bytes: bytes, filename: str) -> RfqDetectionResult:
    """Try to recognize an uploaded file as a natural-stone or brilliant RFQ
    and extract the item(s) it requests.

    Only .xlsx is attempted: both known RFQ layouts (and the only example
    fixtures available) are XLSX. Anything else - or an XLSX that doesn't
    match either layout - is reported as unrecognized so the caller falls
    back to manual material/supplier selection.
    """
    if not filename.lower().endswith(".xlsx"):
        return RfqDetectionResult(recognized=False)

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        return RfqDetectionResult(recognized=False)

    natural_stone_result = _detect_natural_stone_rfq(workbook)
    if natural_stone_result is not None:
        return natural_stone_result

    brilliant_result = _detect_brilliant_rfq(workbook)
    if brilliant_result is not None:
        return brilliant_result

    return RfqDetectionResult(recognized=False)


# ---------------------------------------------------------------------
# Natural-stone RFQ detection (e.g. prirodni.xlsx)
# ---------------------------------------------------------------------

def _normalize_for_matching(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _best_catalog_match(
    description: str, catalog_names: list[str]
) -> tuple[float, str | None]:
    """Fuzzy-match a free-text stone description against catalog goods_name
    entries, trying progressively shorter leading-word windows (stone names
    are always described before shape/size details in the real fixture,
    e.g. "Amethyst afrikan round regular 7 mm")."""
    catalog_by_norm = {_normalize_for_matching(name): name for name in catalog_names}
    tokens = _normalize_for_matching(description).split()

    best_score = 0.0
    best_goods_name: str | None = None

    for window in range(min(4, len(tokens)), 0, -1):
        candidate = " ".join(tokens[:window])
        matches = difflib.get_close_matches(
            candidate, catalog_by_norm.keys(), n=1, cutoff=0.0
        )
        if not matches:
            continue

        ratio = difflib.SequenceMatcher(None, candidate, matches[0]).ratio()
        if ratio > best_score:
            best_score = ratio
            best_goods_name = catalog_by_norm[matches[0]]

    return best_score, best_goods_name


def _detect_natural_stone_rfq(workbook) -> RfqDetectionResult | None:
    catalog = repo.list_goods_names_by_group(NATURAL_STONE_GOODS_GROUP)
    if not catalog:
        return None

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        header = find_header_row(ws, {"description"})
        if header is None:
            continue

        header_row_index, header_cols = header
        description_col = header_cols["description"]
        quantity_col = find_column_containing(ws, header_row_index, "quantity")

        items: list[DetectedRfqItem] = []
        unresolved_lines: list[str] = []

        for row_index in range(header_row_index + 1, ws.max_row + 1):
            description_text = clean_text(
                ws.cell(row=row_index, column=description_col).value
            )
            if not description_text:
                continue

            quantity_value = None
            if quantity_col is not None:
                quantity_value = coerce_float(
                    ws.cell(row=row_index, column=quantity_col).value
                )

            score, goods_name = _best_catalog_match(description_text, catalog)
            if goods_name is not None and score >= NATURAL_STONE_MATCH_THRESHOLD:
                # Each row is kept as its own item, never merged with other
                # rows matching the same catalog stone: different sizes (and
                # shapes) of the same stone can carry a different price per
                # carat, so each becomes its own subcase.
                items.append(
                    DetectedRfqItem(
                        goods_name=goods_name,
                        description=description_text,
                        quantity=quantity_value or None,
                        item_label=description_text,
                    )
                )
            else:
                unresolved_lines.append(description_text)

        if not items:
            continue

        return RfqDetectionResult(
            recognized=True,
            file_type="natural stone",
            items=items,
            unresolved_lines=unresolved_lines,
        )

    return None


# ---------------------------------------------------------------------
# Brilliant RFQ detection (e.g. brilianty.xlsx)
#
# Rows are grouped into lots (see brilliant_lot_parser.py): a lot's several
# component sizes share one merged ORDER CTS TOTAL and one merged Price
# USD/ct cell. One case_item is created per lot, never per row and never
# merged across lots - even when several lots resolve to the same catalog
# bucket (e.g. several lots are all "up to 1ct"), since each lot is priced
# independently by the supplier.
# ---------------------------------------------------------------------

def _detect_color_from_notes(ws, header_row_index: int) -> str | None:
    """Brilliant RFQ color/clarity is stated once for the whole batch in the
    free-text note rows above the size table, not per row."""
    for row_index in range(1, header_row_index):
        for col_index in range(1, ws.max_column + 1):
            text = clean_text(ws.cell(row=row_index, column=col_index).value).lower()
            if not text:
                continue
            for color in _BRILLIANT_COLOR_KEYWORDS:
                if color in text:
                    return color

    return None


def _brilliant_goods_name(color: str | None, bucket: str) -> str:
    threshold_text = "1ct and up" if bucket == "1ct_and_up" else "up to 1ct"
    if color:
        return f"{color.capitalize()} Diamonds {threshold_text}"
    return threshold_text


def _detect_brilliant_rfq(workbook) -> RfqDetectionResult | None:
    catalog = repo.list_goods_names_by_group(BRILLIANT_GOODS_GROUP)
    if not catalog:
        return None

    catalog_by_norm = {name.strip().lower(): name for name in catalog}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        parsed = parse_brilliant_lots(ws)
        if parsed is None:
            continue

        color = _detect_color_from_notes(ws, parsed.header_row_index)

        items: list[DetectedRfqItem] = []
        unresolved_lines: list[str] = []

        for lot in parsed.lots:
            if lot.size_bucket is None:
                unresolved_lines.append(
                    f"{lot.label}: component sizes imply inconsistent "
                    "supplier categories (mixing up-to-1ct and 1ct-and-up)"
                )
                continue

            if not lot.order_cts_total or lot.order_cts_total <= 0:
                unresolved_lines.append(
                    f"{lot.label}: missing or invalid ORDER CTS TOTAL"
                )
                continue

            candidate_name = _brilliant_goods_name(color, lot.size_bucket)
            goods_name = catalog_by_norm.get(candidate_name.lower())
            if goods_name is None:
                unresolved_lines.append(
                    f"{lot.label} (no catalog entry for {candidate_name!r})"
                )
                continue

            items.append(
                DetectedRfqItem(
                    goods_name=goods_name,
                    description=build_lot_description(
                        lot.rows, lot.order_cts_total
                    ),
                    quantity=lot.order_cts_total,
                    item_label=lot.label,
                )
            )

        if not items:
            continue

        return RfqDetectionResult(
            recognized=True,
            file_type="brilliant",
            items=items,
            unresolved_lines=unresolved_lines,
        )

    return None
