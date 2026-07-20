from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import os

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

def get_default_xlsx_path() -> Path:
    configured = os.getenv(
        "SUPPLIER_CATALOG_XLSX",
        "data/supplier_catalog.xlsx",
    )

    path = Path(configured)

    if not path.is_absolute():
        path = PROJECT_ROOT / path

    return path

from app.db.database import get_connection, initialize_database


SUPPLIER_SHEET = "Seznam SUPPLIERS"
FILTER_SHEET = "Supplier Filter"


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def normalize_name(value: str) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = " ".join(text.split())
    return text


def make_slug(value: str, max_len: int = 42) -> str:
    text = clean_text(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", ".", text)
    text = re.sub(r"\.+", ".", text).strip(".")
    if not text:
        text = "supplier"
    return text[:max_len].strip(".") or "supplier"


def make_supplier_code(name: str, used_codes: set[str]) -> str:
    slug = make_slug(name, max_len=28).replace(".", "_").upper()
    code = f"SUP_{slug}"
    code = re.sub(r"_+", "_", code).strip("_")
    if len(code) > 48:
        code = code[:48].rstrip("_")

    base = code
    suffix = 2
    while code in used_codes:
        tail = f"_{suffix}"
        code = f"{base[:48 - len(tail)]}{tail}"
        suffix += 1

    used_codes.add(code)
    return code


def make_dummy_email(name: str) -> str:
    slug = make_slug(name)
    return f"{slug}@{slug}.com"


def make_dummy_whatsapp(index: int) -> str:
    # Deliberately dummy Czech-looking numbers. For simulation only.
    return f"420700{index:06d}"


@dataclass(frozen=True)
class SupplierRow:
    name: str
    contact_person: str | None
    email: str
    whatsapp_number: str
    supplier_code: str


@dataclass(frozen=True)
class MappingRow:
    supplier_filter_name: str
    supplier_name: str
    goods_group: str
    goods_name: str
    source_column: str


def iter_supplier_rows(workbook_path: Path) -> list[SupplierRow]:
    wb = load_workbook(workbook_path, data_only=True)
    if SUPPLIER_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook is missing sheet: {SUPPLIER_SHEET!r}")

    ws = wb[SUPPLIER_SHEET]
    used_codes: set[str] = set()
    suppliers: list[SupplierRow] = []

    seen_names: set[str] = set()
    for row_index in range(2, ws.max_row + 1):
        name = clean_text(ws.cell(row=row_index, column=2).value)
        if not name:
            continue

        name_key = normalize_name(name)
        if name_key in seen_names:
            continue
        seen_names.add(name_key)

        contact_person = clean_text(ws.cell(row=row_index, column=3).value) or None
        email = clean_text(ws.cell(row=row_index, column=4).value)
        phone = clean_text(ws.cell(row=row_index, column=5).value)

        suppliers.append(
            SupplierRow(
                name=name,
                contact_person=contact_person,
                email=email or make_dummy_email(name),
                whatsapp_number=re.sub(r"\D+", "", phone) or make_dummy_whatsapp(len(suppliers) + 1),
                supplier_code=make_supplier_code(name, used_codes),
            )
        )

    return suppliers


def build_supplier_name_matcher(suppliers: Iterable[SupplierRow]) -> dict[str, str]:
    exact: dict[str, str] = {}
    for supplier in suppliers:
        norm = normalize_name(supplier.name)
        if norm:
            exact[norm] = supplier.name
    return exact


def match_supplier_name(filter_name: str, suppliers: list[SupplierRow], exact: dict[str, str]) -> str:
    filter_norm = normalize_name(filter_name)
    if filter_norm in exact:
        return exact[filter_norm]

    # Common workbook pattern: Supplier Filter has the short name, while
    # Seznam SUPPLIERS contains the same name plus a parenthetical alias.
    candidates: list[tuple[int, str]] = []
    for supplier in suppliers:
        supplier_norm = normalize_name(supplier.name)
        if not supplier_norm:
            continue
        if supplier_norm.startswith(filter_norm) or filter_norm.startswith(supplier_norm):
            candidates.append((abs(len(supplier_norm) - len(filter_norm)), supplier.name))
        elif filter_norm in supplier_norm or supplier_norm in filter_norm:
            candidates.append((abs(len(supplier_norm) - len(filter_norm)) + 100, supplier.name))

    if candidates:
        candidates.sort(key=lambda item: item[0])
        return candidates[0][1]

    # If the supplier exists only on the filter sheet, preserve it instead
    # of dropping its mapping.
    return clean_text(filter_name)


def excel_column_name(column_index: int) -> str:
    name = ""
    while column_index:
        column_index, rem = divmod(column_index - 1, 26)
        name = chr(65 + rem) + name
    return name


def iter_mapping_rows(workbook_path: Path, suppliers: list[SupplierRow]) -> list[MappingRow]:
    wb = load_workbook(workbook_path, data_only=True)
    if FILTER_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook is missing sheet: {FILTER_SHEET!r}")

    ws = wb[FILTER_SHEET]
    exact = build_supplier_name_matcher(suppliers)

    goods_by_col: dict[int, tuple[str, str]] = {}
    current_group = ""
    for col in range(3, ws.max_column + 1):
        group = clean_text(ws.cell(row=1, column=col).value)
        if group:
            current_group = group

        goods_name = clean_text(ws.cell(row=2, column=col).value)
        if goods_name:
            goods_by_col[col] = (current_group, goods_name)

    mappings: list[MappingRow] = []
    for row in range(3, ws.max_row + 1):
        filter_supplier_name = clean_text(ws.cell(row=row, column=2).value)
        if not filter_supplier_name:
            continue

        supplier_name = match_supplier_name(filter_supplier_name, suppliers, exact)

        for col, (goods_group, goods_name) in goods_by_col.items():
            marker = clean_text(ws.cell(row=row, column=col).value).lower()
            if marker == "x":
                mappings.append(
                    MappingRow(
                        supplier_filter_name=filter_supplier_name,
                        supplier_name=supplier_name,
                        goods_group=goods_group,
                        goods_name=goods_name,
                        source_column=excel_column_name(col),
                    )
                )

    return mappings


def ensure_supplier_goods_schema() -> None:
    with get_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS supplier_goods (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_id INTEGER NOT NULL,
                goods_name TEXT NOT NULL,
                goods_group TEXT,
                source_sheet TEXT NOT NULL DEFAULT 'Supplier Filter',
                source_column TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(supplier_id, goods_name),
                FOREIGN KEY(supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_supplier_goods_goods_name
            ON supplier_goods(goods_name);

            CREATE INDEX IF NOT EXISTS idx_supplier_goods_supplier_id
            ON supplier_goods(supplier_id);
            """
        )
        conn.commit()


def upsert_suppliers_and_mappings(
    suppliers: list[SupplierRow],
    mappings: list[MappingRow],
    *,
    default_channel: str,
    replace_mappings: bool,
    dry_run: bool,
) -> dict[str, int]:
    if default_channel not in {"email", "whatsapp"}:
        raise ValueError("default_channel must be 'email' or 'whatsapp'.")

    existing_supplier_names = {normalize_name(s.name) for s in suppliers}
    extra_supplier_names = sorted(
        {
            mapping.supplier_name
            for mapping in mappings
            if normalize_name(mapping.supplier_name) not in existing_supplier_names
        },
        key=str.lower,
    )

    all_suppliers = list(suppliers)
    used_codes = {supplier.supplier_code for supplier in suppliers}
    for extra_name in extra_supplier_names:
        all_suppliers.append(
            SupplierRow(
                name=extra_name,
                contact_person=None,
                email=make_dummy_email(extra_name),
                whatsapp_number=make_dummy_whatsapp(len(all_suppliers) + 1),
                supplier_code=make_supplier_code(extra_name, used_codes),
            )
        )

    if dry_run:
        return {
            "supplier_rows_from_list": len(suppliers),
            "extra_suppliers_from_filter": len(extra_supplier_names),
            "suppliers_to_upsert": len(all_suppliers),
            "mappings_to_import": len(mappings),
            "goods_categories": len({m.goods_name for m in mappings}),
            "inserted_or_updated_suppliers": 0,
            "inserted_mappings": 0,
        }

    ensure_supplier_goods_schema()

    supplier_ids_by_norm: dict[str, int] = {}
    with get_connection() as conn:
        for supplier in all_suppliers:
            notes_parts = ["Imported from buyer supplier XLSX."]
            if supplier.contact_person:
                notes_parts.append(f"Contact person: {supplier.contact_person}.")
            notes_parts.append("Email and WhatsApp may be dummy presentation data.")
            notes = " ".join(notes_parts)

            existing = conn.execute(
                """
                SELECT id, supplier_code
                FROM suppliers
                WHERE supplier_code = ?
                   OR lower(name) = lower(?)
                ORDER BY CASE WHEN supplier_code = ? THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (supplier.supplier_code, supplier.name, supplier.supplier_code),
            ).fetchone()

            if existing:
                supplier_id = int(existing["id"])
                conn.execute(
                    """
                    UPDATE suppliers
                    SET name = ?,
                        contact_channel = ?,
                        whatsapp_number = ?,
                        email = ?,
                        category = ?,
                        active = 1,
                        notes = ?
                    WHERE id = ?
                    """,
                    (
                        supplier.name,
                        default_channel,
                        supplier.whatsapp_number,
                        supplier.email,
                        "Imported buyer supplier list",
                        notes,
                        supplier_id,
                    ),
                )
            else:
                cur = conn.execute(
                    """
                    INSERT INTO suppliers
                    (supplier_code, name, contact_channel, whatsapp_number, email, category, active, notes)
                    VALUES (?, ?, ?, ?, ?, ?, 1, ?)
                    """,
                    (
                        supplier.supplier_code,
                        supplier.name,
                        default_channel,
                        supplier.whatsapp_number,
                        supplier.email,
                        "Imported buyer supplier list",
                        notes,
                    ),
                )
                supplier_id = int(cur.lastrowid)

            supplier_ids_by_norm[normalize_name(supplier.name)] = supplier_id

        if replace_mappings:
            conn.execute("DELETE FROM supplier_goods")

        inserted_mappings = 0
        for mapping in mappings:
            supplier_id = supplier_ids_by_norm.get(normalize_name(mapping.supplier_name))
            if supplier_id is None:
                raise RuntimeError(f"Internal error: supplier not upserted: {mapping.supplier_name}")

            cur = conn.execute(
                """
                INSERT OR IGNORE INTO supplier_goods
                (supplier_id, goods_name, goods_group, source_sheet, source_column)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    supplier_id,
                    mapping.goods_name,
                    mapping.goods_group or None,
                    FILTER_SHEET,
                    mapping.source_column,
                ),
            )
            inserted_mappings += cur.rowcount

        conn.commit()

    return {
        "supplier_rows_from_list": len(suppliers),
        "extra_suppliers_from_filter": len(extra_supplier_names),
        "suppliers_to_upsert": len(all_suppliers),
        "mappings_to_import": len(mappings),
        "goods_categories": len({m.goods_name for m in mappings}),
        "inserted_or_updated_suppliers": len(all_suppliers),
        "inserted_mappings": inserted_mappings,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import buyer supplier XLSX into the purchasing assistant database."
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        type=Path,
        help="Path to the buyer supplier XLSX file.",
    )
    parser.add_argument(
        "--default-channel",
        choices=("email", "whatsapp"),
        default="email",
        help="Default contact channel assigned to imported suppliers.",
    )
    parser.add_argument(
        "--append-mappings",
        action="store_true",
        help="Append mappings instead of replacing the supplier_goods table.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and print counts without writing to the database.",
    )
    return parser.parse_args()

def get_current_workbook_supplier_names(
    suppliers: list[SupplierRow],
    mappings: list[MappingRow],
) -> list[str]:
    supplier_names = {
        supplier.name
        for supplier in suppliers
    }

    supplier_names.update(
        mapping.supplier_name
        for mapping in mappings
    )

    return sorted(supplier_names, key=str.lower)


def deactivate_suppliers_not_in_workbook(
    current_supplier_names: list[str],
) -> int:
    """
    The XLSX is now the active supplier catalog.

    We deactivate suppliers not present in the workbook instead of deleting
    them, because old cases/messages/offers may still reference them.
    """
    if not current_supplier_names:
        return 0

    normalized_names = [
        name.lower()
        for name in current_supplier_names
    ]

    placeholders = ", ".join("?" for _ in normalized_names)

    with get_connection() as conn:
        row = conn.execute(
            f"""
            SELECT COUNT(*) AS count_to_deactivate
            FROM suppliers
            WHERE active = 1
              AND lower(name) NOT IN ({placeholders})
            """,
            tuple(normalized_names),
        ).fetchone()

        count_to_deactivate = int(row["count_to_deactivate"])

        conn.execute(
            f"""
            UPDATE suppliers
            SET active = 0
            WHERE active = 1
              AND lower(name) NOT IN ({placeholders})
            """,
            tuple(normalized_names),
        )

        conn.commit()

    return count_to_deactivate


def import_workbook(
    xlsx_path: Path | None = None,
    dry_run: bool = False,
    default_channel: str = "email",
    replace_mappings: bool = True,
) -> dict[str, int]:
    """
    Import the buyer XLSX into suppliers and supplier_goods.

    This function is used by:
    - scripts/import_supplier_filter_xlsx.py
    - scripts/init_db.py

    It intentionally reuses the parser functions already present in this file:
    - iter_supplier_rows(...)
    - iter_mapping_rows(...)
    - upsert_suppliers_and_mappings(...)
    """
    xlsx_path = xlsx_path or get_default_xlsx_path()

    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    suppliers = iter_supplier_rows(xlsx_path)
    mappings = iter_mapping_rows(xlsx_path, suppliers)

    current_supplier_names = get_current_workbook_supplier_names(
        suppliers=suppliers,
        mappings=mappings,
    )

    result = upsert_suppliers_and_mappings(
        suppliers=suppliers,
        mappings=mappings,
        default_channel=default_channel,
        replace_mappings=replace_mappings,
        dry_run=dry_run,
    )

    if dry_run:
        result["suppliers_to_deactivate"] = 0
        return result

    result["suppliers_to_deactivate"] = deactivate_suppliers_not_in_workbook(
        current_supplier_names
    )

    return result

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Import buyer XLSX supplier/material filter into "
            "the purchasing assistant database."
        )
    )

    parser.add_argument(
        "--xlsx",
        required=False,
        type=Path,
        help=(
            "Path to the buyer XLSX file. "
            "Defaults to SUPPLIER_CATALOG_XLSX or data/supplier_catalog.xlsx."
        ),
    )

    parser.add_argument(
        "--default-channel",
        choices=("email", "whatsapp"),
        default="email",
        help="Default contact channel assigned to imported suppliers.",
    )

    parser.add_argument(
        "--append-mappings",
        action="store_true",
        help="Append mappings instead of replacing the supplier_goods table.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and print counts without writing to the database.",
    )

    args = parser.parse_args()

    xlsx_path = args.xlsx or get_default_xlsx_path()

    summary = import_workbook(
        xlsx_path=xlsx_path,
        dry_run=args.dry_run,
        default_channel=args.default_channel,
        replace_mappings=not args.append_mappings,
    )

    print("DRY RUN" if args.dry_run else "IMPORT COMPLETE")
    print(f"Workbook: {xlsx_path}")

    for key, value in summary.items():
        print(f"{key}: {value}")

    if not args.dry_run:
        print("Supplier catalog is now synchronized from XLSX.")
        print("Old suppliers not present in the XLSX were deactivated, not deleted.")

if __name__ == "__main__":
    main()
