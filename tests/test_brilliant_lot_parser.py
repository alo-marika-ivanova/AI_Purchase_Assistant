from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.llm.rfq_multi_item_price_safeguard import extract_safe_multi_item_rfq_prices
from app.services.brilliant_lot_parser import (
    parse_brilliant_lots,
    try_build_brilliant_lot_offer_text,
)

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures" / "rfq_samples"

EXPECTED_LOT_TOTALS = [350.0, 200.0, 160.0, 250.0, 100.0, 100.0, 100.0]
EXPECTED_LOT_PRICES = [42.5, 55.0, 61.25, 70.0, 95.5, 110.25, 130.75]


def _read_fixture_bytes(name: str) -> bytes:
    return (FIXTURES_DIR / name).read_bytes()


def _load_sheet(name: str):
    workbook = load_workbook(FIXTURES_DIR / name, data_only=True)
    return workbook[workbook.sheetnames[0]]


def test_parse_brilliant_lots_finds_seven_lots_in_blank_rfq_template() -> None:
    parsed = parse_brilliant_lots(_load_sheet("brilianty.xlsx"))

    assert parsed is not None
    assert len(parsed.lots) == 7
    assert [lot.order_cts_total for lot in parsed.lots] == EXPECTED_LOT_TOTALS
    assert all(lot.price_usd_per_ct is None for lot in parsed.lots)
    assert all(lot.size_bucket == "up_to_1ct" for lot in parsed.lots)


def test_parse_brilliant_lots_first_lot_label_matches_expected_mm_range() -> None:
    parsed = parse_brilliant_lots(_load_sheet("brilianty.xlsx"))

    assert parsed.lots[0].label == "Brilliant lot 1: 0.66–1.25 mm"


def test_parse_brilliant_lots_mixed_mm_and_ct_lot_gets_a_dual_range_label() -> None:
    parsed = parse_brilliant_lots(_load_sheet("brilianty.xlsx"))

    # Lot 4 contains both mm rows (2.71-3.00mm) and ct rows (0.11-0.13ct);
    # the label must show both ranges rather than guess a conversion.
    assert (
        parsed.lots[3].label
        == "Brilliant lot 4: 2.71–3.00 mm / 0.11–0.13 ct"
    )


def test_parse_brilliant_lots_reads_filled_prices() -> None:
    parsed = parse_brilliant_lots(_load_sheet("brilianty_filled.xlsx"))

    assert [lot.price_usd_per_ct for lot in parsed.lots] == EXPECTED_LOT_PRICES


def test_parse_brilliant_lots_returns_none_for_natural_stone_workbook() -> None:
    assert parse_brilliant_lots(_load_sheet("prirodni.xlsx")) is None


def test_try_build_brilliant_lot_offer_text_prices_every_lot() -> None:
    text = try_build_brilliant_lot_offer_text(
        _read_fixture_bytes("brilianty_filled.xlsx"), "brilianty_filled.xlsx"
    )

    assert text is not None
    lines = text.splitlines()
    assert lines[0] == "Description | Price USD/ct"
    assert len(lines) == 8

    parsed = parse_brilliant_lots(_load_sheet("brilianty_filled.xlsx"))
    for lot in parsed.lots:
        assert f"{lot.label} | {lot.price_usd_per_ct}" in lines


def test_try_build_brilliant_lot_offer_text_returns_none_for_blank_rfq_template() -> None:
    """The outgoing (unfilled) RFQ has no prices at all - it must not be
    mistaken for a completed reply."""
    text = try_build_brilliant_lot_offer_text(
        _read_fixture_bytes("brilianty.xlsx"), "brilianty.xlsx"
    )

    assert text is None


def test_try_build_brilliant_lot_offer_text_returns_none_when_one_lot_price_is_missing(
    tmp_path,
) -> None:
    """A partially filled reply (one lot's price left blank) must not be
    guessed at, and the missing price must not leak into another lot."""
    workbook = load_workbook(FIXTURES_DIR / "brilianty_filled.xlsx", data_only=False)
    worksheet = workbook["List2"]
    worksheet.cell(row=45, column=5).value = None  # blank out lot 5's price

    partial_path = tmp_path / "brilianty_partial.xlsx"
    workbook.save(partial_path)

    text = try_build_brilliant_lot_offer_text(
        partial_path.read_bytes(), "brilianty_partial.xlsx"
    )

    assert text is None


def test_try_build_brilliant_lot_offer_text_returns_none_for_unrelated_workbook() -> None:
    """A natural-stone reply must keep using the generic flattened dump,
    not be swallowed by the brilliant-lot parser."""
    text = try_build_brilliant_lot_offer_text(
        _read_fixture_bytes("prirodni.xlsx"), "prirodni.xlsx"
    )

    assert text is None


def _lot_labels(name: str) -> list[str]:
    parsed = parse_brilliant_lots(_load_sheet(name))
    return [lot.label for lot in parsed.lots]


def test_deterministic_multi_item_safeguard_prices_all_seven_lots() -> None:
    """The brilliant-lot reply text plugs straight into the EXISTING
    multi-item deterministic safeguard (used today for natural-stone
    replies) with no change to that module: matching is purely by exact
    item_material text against each lot's stable label."""
    labels = _lot_labels("brilianty_filled.xlsx")
    supplier_items = [
        {"item_material": label, "id": index} for index, label in enumerate(labels)
    ]

    text = try_build_brilliant_lot_offer_text(
        _read_fixture_bytes("brilianty_filled.xlsx"), "brilianty_filled.xlsx"
    )

    offers = extract_safe_multi_item_rfq_prices(
        message_body=text, supplier_items=supplier_items
    )

    assert offers is not None
    prices_by_material = {
        offer["item_material"]: offer["unit_price_usd"] for offer in offers
    }
    for label, expected_price in zip(labels, EXPECTED_LOT_PRICES):
        assert prices_by_material[label] == expected_price
        assert (
            next(o for o in offers if o["item_material"] == label)["price_certainty"]
            == "CONFIRMED"
        )


def test_deterministic_multi_item_safeguard_bails_when_a_required_lot_is_absent() -> None:
    """If a lot this supplier was asked to quote never shows up in the
    reply text at all, the safeguard must bail rather than leave that
    item unpriced while silently accepting the rest."""
    labels = _lot_labels("brilianty_filled.xlsx")
    supplier_items = [
        {"item_material": label, "id": index} for index, label in enumerate(labels)
    ]

    lines = ["Description | Price USD/ct"]
    for label, price in zip(labels[:-1], EXPECTED_LOT_PRICES[:-1]):
        lines.append(f"{label} | {price}")
    partial_text = "\n".join(lines)

    offers = extract_safe_multi_item_rfq_prices(
        message_body=partial_text, supplier_items=supplier_items
    )

    assert offers is None
