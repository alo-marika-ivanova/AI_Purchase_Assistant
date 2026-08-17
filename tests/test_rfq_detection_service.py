from __future__ import annotations

from pathlib import Path

from app.db.database import get_connection
from app.services.rfq_detection_service import detect_rfq_selection


FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures" / "rfq_samples"

NATURAL_STONE_GOODS_NAMES = [
    "Amethyst African (AMA)",
    "Amethyst Green (AMG)",
    "Citrine Madeira (CIM)",
    "Garnet Pink",
    "Peridote (PER)",
    "Prehnite (PRE)",
    "Rhodolite Purple (RHF)",
    "Rhodolite Reddish (RHO)",
    "Spinel (SPI)",
    "Tanzanite (TAN)",
    "Topaz London (TOL)",
    "Topaz Sky (TOY)",
    "Topaz Swiss (TOS)",
]


def _seed_supplier_goods(
    supplier_id: int, goods_group: str, goods_names: list[str]
) -> None:
    with get_connection() as conn:
        conn.executemany(
            """
            INSERT INTO supplier_goods (supplier_id, goods_name, goods_group)
            VALUES (?, ?, ?)
            """,
            [(supplier_id, goods_name, goods_group) for goods_name in goods_names],
        )
        conn.commit()


def _read_fixture(name: str) -> bytes:
    return (FIXTURES_DIR / name).read_bytes()


def test_natural_stone_rfq_is_recognized_and_matched(
    supplier_ids: dict[str, int],
) -> None:
    _seed_supplier_goods(
        supplier_ids["email"], "Precious stones", NATURAL_STONE_GOODS_NAMES
    )

    result = detect_rfq_selection(_read_fixture("prirodni.xlsx"), "prirodni.xlsx")

    assert result.recognized is True
    assert result.file_type == "natural stone"
    assert result.unresolved_lines == []

    detected_names = {item.goods_name for item in result.items}
    assert detected_names == set(NATURAL_STONE_GOODS_NAMES)


def test_natural_stone_rfq_keeps_each_row_as_its_own_item(
    supplier_ids: dict[str, int],
) -> None:
    """Rows matching the same catalog stone are never merged: different
    sizes/shapes of the same stone can carry a different price per carat,
    so each source row becomes its own item (and later its own subcase)."""
    _seed_supplier_goods(
        supplier_ids["email"], "Precious stones", NATURAL_STONE_GOODS_NAMES
    )

    result = detect_rfq_selection(_read_fixture("prirodni.xlsx"), "prirodni.xlsx")

    amethyst_items = [
        item for item in result.items if item.goods_name == "Amethyst African (AMA)"
    ]
    assert sorted(item.quantity for item in amethyst_items) == [12.0, 16.0, 16.0, 16.0]
    assert {item.description for item in amethyst_items} == {
        "Amethyst afrikan round regular 7 mm",
        "Amethyst afrikan cushion regular 6 mm",
        "Amethyst African octagon regular 9x7 mm",
        "Amethyst African oval regular 5x4 mm",
    }

    peridote_items = [
        item for item in result.items if item.goods_name == "Peridote (PER)"
    ]
    assert sorted(item.quantity for item in peridote_items) == [16.0, 24.0, 100.0]
    assert {item.description for item in peridote_items} == {
        "Peridot round regular 2 mm",
        "Peridot round regular 4 mm",
        "Peridot round regular 5 mm",
    }

    # Each item's display_name (used as the subcase's item_material) carries
    # the distinguishing per-row text, not just the shared catalog name.
    assert {item.display_name for item in peridote_items} == {
        "Peridot round regular 2 mm",
        "Peridot round regular 4 mm",
        "Peridot round regular 5 mm",
    }


def test_natural_stone_rfq_reports_unresolved_lines_for_unmatched_stones(
    supplier_ids: dict[str, int],
) -> None:
    # Only seed a subset of the catalog - the rest should fall to unresolved.
    _seed_supplier_goods(
        supplier_ids["email"], "Precious stones", ["Amethyst African (AMA)"]
    )

    result = detect_rfq_selection(_read_fixture("prirodni.xlsx"), "prirodni.xlsx")

    assert result.recognized is True
    detected_names = {item.goods_name for item in result.items}
    assert detected_names == {"Amethyst African (AMA)"}
    assert len(result.unresolved_lines) > 0
    assert any("Tanzanite" in line for line in result.unresolved_lines)


def test_brilliant_rfq_is_recognized_as_seven_separate_lot_items(
    supplier_ids: dict[str, int],
) -> None:
    """Rows are grouped into lots (a merged ORDER CTS TOTAL / Price USD/ct
    per lot), not into one item for the whole workbook and not into one
    item per individual size row: brilianty.xlsx has 7 lots totalling
    350+200+160+250+100+100+100 = 1260 ct."""
    _seed_supplier_goods(
        supplier_ids["email"], "Diamonds", ["up to 1ct", "1ct and up"]
    )

    result = detect_rfq_selection(_read_fixture("brilianty.xlsx"), "brilianty.xlsx")

    assert result.recognized is True
    assert result.file_type == "brilliant"
    assert result.unresolved_lines == []
    assert len(result.items) == 7
    assert [item.quantity for item in result.items] == [
        350.0, 200.0, 160.0, 250.0, 100.0, 100.0, 100.0,
    ]

    # All 7 lots resolve to the same colorless "up to 1ct" catalog bucket
    # (used for supplier selection) but must remain 7 separate items, since
    # each lot is priced independently by the supplier.
    assert {item.goods_name for item in result.items} == {"up to 1ct"}
    assert len({item.display_name for item in result.items}) == 7


def test_brilliant_rfq_first_lot_has_the_expected_stable_label(
    supplier_ids: dict[str, int],
) -> None:
    _seed_supplier_goods(
        supplier_ids["email"], "Diamonds", ["up to 1ct", "1ct and up"]
    )

    result = detect_rfq_selection(_read_fixture("brilianty.xlsx"), "brilianty.xlsx")

    assert result.items[0].display_name == "Brilliant lot 1: 0.66–1.25 mm"
    assert "350" in result.items[0].description


def test_detection_falls_back_when_catalog_is_empty(
    supplier_ids: dict[str, int],
) -> None:
    result = detect_rfq_selection(_read_fixture("prirodni.xlsx"), "prirodni.xlsx")

    assert result.recognized is False
    assert result.items == []


def test_detection_ignores_non_xlsx_filenames(
    supplier_ids: dict[str, int],
) -> None:
    result = detect_rfq_selection(b"item,quantity\nfoo,1\n", "offer.csv")

    assert result.recognized is False


def test_detection_reports_unrecognized_for_unrelated_xlsx(
    supplier_ids: dict[str, int],
) -> None:
    _seed_supplier_goods(
        supplier_ids["email"], "Precious stones", NATURAL_STONE_GOODS_NAMES
    )

    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    ws = workbook.active
    ws["A1"] = "Not an RFQ"
    ws["A2"] = "Just some notes"
    buffer = BytesIO()
    workbook.save(buffer)

    result = detect_rfq_selection(buffer.getvalue(), "notes.xlsx")

    assert result.recognized is False
